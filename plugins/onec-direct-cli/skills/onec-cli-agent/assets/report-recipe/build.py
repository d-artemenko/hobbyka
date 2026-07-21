#!/usr/bin/env python3
import argparse
import hashlib
import json
import os
import re
import subprocess
from pathlib import Path

from openpyxl import Workbook


def run_query(cli, path):
    result = subprocess.run(
        ["node", cli, "query", "--file", str(path), "--max-rows", "2000"],
        text=True,
        capture_output=True,
        check=True,
    )
    payload = json.loads(result.stdout)
    if not payload.get("ok"):
        raise RuntimeError(payload.get("error", f"1C query failed: {path.name}"))
    data = payload["data"]
    assert data["rowCount"] == len(data["rows"]), path.name
    assert data["columns"], path.name
    return data


def sheet_name(name, used):
    base = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "data"
    candidate = base
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--cli", default=os.environ.get("ONEC_CLI"))
    parser.add_argument("--output", default="artifacts/report.xlsx")
    args = parser.parse_args()
    if not args.cli:
        parser.error("pass --cli or set ONEC_CLI")

    root = Path(__file__).resolve().parent
    queries = sorted(root.glob("*.q1c"))
    if not queries:
        parser.error("add at least one .q1c file next to build.py")

    workbook = Workbook()
    workbook.remove(workbook.active)
    sources = workbook.create_sheet("Источники")
    sources.append(["Запрос", "Строк", "SHA-256"])
    used = {"Источники"}

    for query in queries:
        data = run_query(args.cli, query)
        worksheet = workbook.create_sheet(sheet_name(query.stem, used))
        worksheet.append(data["columns"])
        for row in data["rows"]:
            worksheet.append(row)
        digest = hashlib.sha256(query.read_bytes()).hexdigest()
        sources.append([query.name, data["rowCount"], digest])

    output = Path(args.output)
    if not output.is_absolute():
        output = root / output
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(output.resolve())


if __name__ == "__main__":
    main()
